import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill
import re

def clean_amount(amount_str):
    """Extract and clean amount from string, handling different formats."""
    if pd.isna(amount_str):
        return None
    
    # Convert to string if not already
    amount_str = str(amount_str)
    
    # Remove currency symbols and handle negative amounts
    amount_str = amount_str.replace('£', '').replace(',', '')
    
    # Extract amount using regex
    match = re.search(r'-?\d+\.?\d*', amount_str)
    if match:
        return float(match.group())
    return None

def parse_date(date_str):
    """Parse dates from different formats."""
    if pd.isna(date_str) or str(date_str).lower() == 'pending':
        return None
    
    try:
        # Try parsing with various formats
        for fmt in ['%Y-%m-%d %H:%M:%S %z', '%d/%m/%Y', '%Y-%m-%d']:
            try:
                dt = pd.to_datetime(date_str, format=fmt, dayfirst=True)
                # Convert to UTC if timezone aware, then remove timezone info
                if dt.tzinfo is not None:
                    dt = dt.tz_convert('UTC').tz_localize(None)
                return dt
            except:
                continue
                
        # If no format matches, try general parsing
        dt = pd.to_datetime(date_str, dayfirst=True)
        if dt.tzinfo is not None:
            dt = dt.tz_convert('UTC').tz_localize(None)
        return dt
    except:
        print(f"Could not parse date: {date_str}")
        return None

def convert_to_utc(dt):
    """Convert datetime to UTC and ensure it's timezone-aware."""
    if dt is None:
        return None
    try:
        if dt.tzinfo is None:
            return dt.tz_localize('Europe/London').tz_convert('UTC')
        return dt.tz_convert('UTC')
    except:
        return dt

def load_amazon_data(filename):
    df = pd.read_csv(filename)
    df['source'] = 'Amazon'
    df['date'] = df['date'].apply(parse_date)
    df['amount'] = df['total'].apply(clean_amount)
    df['description'] = df['items']
    return df[['date', 'amount', 'description', 'source']]

def load_paypal_data(filename):
    df = pd.read_csv(filename)
    # Filter out pending transactions and deposits
    df = df[
        (df['Status'] == 'Completed') & 
        (df['Type'] != 'Bank deposit to PayPal account')
    ]
    df['source'] = 'PayPal'
    df['date'] = df['Date'].apply(parse_date)
    df['amount'] = df['Amount'].apply(clean_amount)
    df['description'] = df['Name']
    return df[['date', 'amount', 'description', 'source']]

def load_bank_data(filename):
    # Read CSV with more flexible parsing
    df = pd.read_csv(filename, on_bad_lines='skip')
    df['source'] = 'Bank'
    df['date'] = df['Date'].apply(parse_date)
    df['amount'] = df['Amount'].apply(clean_amount)
    df['description'] = df['Description']
    return df[['date', 'amount', 'description', 'source']]

def load_gmail_data(filename):
    df = pd.read_csv(filename)
    df['source'] = 'Gmail'
    df['date'] = df['Date'].apply(parse_date)
    df['amount'] = df['Amount'].apply(clean_amount)
    df['description'] = df['Subject'] + ' - ' + df['Supplier']
    return df[['date', 'amount', 'description', 'source']]

def find_duplicates(df):
    """Find potential duplicates based on amount within 7 days."""
    duplicates = []
    colors = ['FFB6C1', 'AFEEEE', '98FB98', 'DDA0DD', 'F0E68C', 'E6E6FA', 'FFB347']  # Light colors
    color_index = 0
    
    # Sort by date
    df = df.sort_values('date')
    
    # Only look at rows with non-null amounts
    df_with_amounts = df[df['amount'].notna()].copy()
    
    processed_indices = set()
    
    for i, row in df_with_amounts.iterrows():
        if i in processed_indices:
            continue
            
        amount = row['amount']
        date = row['date']
        
        # Find all transactions with the same amount within 7 days
        mask = (
            (df_with_amounts['amount'] == amount) &
            (df_with_amounts.index != i) &
            (df_with_amounts['date'] - date <= timedelta(days=7)) &
            (df_with_amounts['date'] - date >= timedelta(days=-7))
        )
        
        matching_rows = df_with_amounts[mask]
        
        if len(matching_rows) > 0:
            # Add the current row and all matching rows
            group = pd.concat([pd.DataFrame([row]), matching_rows])
            duplicates.append({
                'indices': group.index.tolist(),
                'color': colors[color_index % len(colors)]
            })
            
            # Mark all these indices as processed
            processed_indices.update(group.index.tolist())
            
            color_index += 1
    
    return duplicates

def save_to_excel(df, duplicates, output_filename):
    """Save to Excel with color-coded duplicates."""
    # Create a copy of the dataframe to avoid modifying the original
    df_excel = df.copy()
    
    # Remove timezone information from dates
    df_excel['date'] = df_excel['date'].apply(lambda x: x.tz_localize(None) if x is not None and x.tzinfo is not None else x)
    
    # Create Excel writer
    writer = pd.ExcelWriter(output_filename, engine='openpyxl')
    
    # Convert to Excel
    df_excel.to_excel(writer, index=False, sheet_name='Transactions')
    
    # Get the workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Transactions']
    
    # Apply colors to duplicate groups
    for dup in duplicates:
        color = dup['color']
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        
        for idx in dup['indices']:
            row_number = df_excel.index.get_loc(idx) + 2  # +2 for header and 1-based indexing
            for col in range(1, len(df_excel.columns) + 1):
                cell = worksheet.cell(row=row_number, column=col)
                cell.fill = fill
    
    # Save the file
    writer.close()

def main():
    # Load data from all sources
    amazon_df = load_amazon_data('amazon_order_history.csv')
    paypal_df = load_paypal_data('paypal.csv')
    bank_df = load_bank_data('bank.csv')
    gmail_df = load_gmail_data('gmail_transactions.csv')
    
    # Combine all dataframes
    combined_df = pd.concat([amazon_df, paypal_df, bank_df, gmail_df], ignore_index=True)
    
    # Sort by date
    combined_df = combined_df.sort_values('date')
    
    # Find potential duplicates
    duplicates = find_duplicates(combined_df)
    
    # Save to Excel with color coding
    save_to_excel(combined_df, duplicates, 'consolidated_transactions.xlsx')
    
    # Print summary
    print(f"Processed transactions:")
    print(f"Amazon: {len(amazon_df)} transactions")
    print(f"PayPal: {len(paypal_df)} transactions")
    print(f"Bank: {len(bank_df)} transactions")
    print(f"Gmail: {len(gmail_df)} transactions")
    print(f"\nTotal: {len(combined_df)} transactions")
    print(f"Found {len(duplicates)} groups of potential duplicates")
    print("\nResults saved to 'consolidated_transactions.xlsx'")

if __name__ == '__main__':
    main()
